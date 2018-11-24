from collections import Counter
from Stop_Words import *
from Curse_Words import *
import Directory_Navigator as DN
import Web_Parser as web_parser
import json
import logging
import os
import openpyxl


def raw_word_count(words):
    """Creates a dictionary of words and their frequencies.
       This will count all words
    Args:    words list
    Returns: sorted dictionary of words and thier frequencies
    """

    words_counter = Counter()
    words_counter.update(words.split())
    words_counter.most_common()
    word_dictionary = dict(words_counter)

    sorted_dictionary = dict(sorted(word_dictionary.items(), key=lambda kv: kv[1], reverse=True))

    return sorted_dictionary


def word_count(words, have_stop_words=False, have_curse_words=True):
    """Returns dictionary of words and their frequencies.
		Have the option to remove stop words and/or curse words.
       Args:    word_list, stop_words boolean, curse words boolean
       Returns: sorted_dictionary of counted words
    """

    words_counter = Counter()
    words = words.split()

    word_list_counter = []
    for item in words:

        if have_stop_words and have_curse_words:
            # Want both stop words and curse words included
            word_list_counter.append(item)

        elif have_stop_words and not have_curse_words:
            # Want just stop words included but curse words removed
            if item not in curse_words:
                word_list_counter.append(item)

        elif not have_stop_words and have_curse_words:
            # Want stop words removed but curse words included
            if item not in stop_words:
                word_list_counter.append(item)

        else:
            # Want to remove both stop and curse words
            if item not in stop_words and item not in curse_words:
                word_list_counter.append(item)

    words_counter.update(word_list_counter)
    words_counter.most_common()
    word_dictionary = dict(words_counter)

    sorted_dictionary = dict(sorted(word_dictionary.items(), key=lambda kv: kv[1], reverse=True))

    return sorted_dictionary


def count_text_file_words(file_path, have_stop_words, have_curse_words):
    """Counts the words in a text file to make word cloud
       Args:        path to text file, have stop words boolean, have curse words boolean
       Returns:     dictionary of words and their frequencies 
    """
    try:
        with open(file_path) as text_file:
            data = text_file.read()

        cleaned_data = web_parser.clean_lyrics(data)
        word_frequency_dict = word_count(cleaned_data, have_stop_words, have_curse_words)


        return word_frequency_dict

    except Exception as e:
        return None

def write_word_count_json(word_dictionary, raw_lyrics=False):
    """Returns a filename of a json file whose most frequently used words are stored.
       Args:    word_dictionary, raw_lyrics boolean
       Returns: filename  
    """
    try:

        file_out = DN.write_json(raw_lyrics, word_dictionary)

        return file_out

    except Exception as e:
        # if word doesnt exit or error return none
        print(e)
        logging.basicConfig(filename='Cloud_lyrics.log', filemode='a', format='%(asctime)s - %(message)s',
                            datefmt='%b-%d-%Y %H:%M:%S')
        logging.error('Error writing word count to json file')
        logging.error('------------------')
        logging.error("Exception occurred", exc_info=True)
        logging.error('------------------')


def write_word_count_excel(word_dictionary_raw, genre, song_artist_list=None):
    """Returns a file directory of the excel file holding the word and used count
       Args:    raw word count dictionary, genre running wordcloud on
       Returns: file path of new excel file
    """
    file_out = DN.write_excel(word_dictionary_raw, genre, song_artist_list)

    return file_out

def create_word_dictionary_from_excel(excel_file_path, have_stop_words=False, have_curse_words=True):
    """Reads an exisiting excel file and turns the words and the frequencies into a dictionary
       Args:    excel_file_path, have_stop_words, have_curse_words
       Returns: A dictionary of words and their frequencies
    """

    try:

        word_count_dictionary = {}

        work_book = openpyxl.load_workbook(excel_file_path) 
          
        # Get workbook active sheet object 
        # from the active attribute 
        sheet = work_book['Word Count']
        row_count = sheet.max_row

        for i in range(2,row_count):
            #print(sheet.cell(row=i, column=1).value)
            word = sheet.cell(row=i, column=1).value
            count = sheet.cell(row=i, column=2).value

            if have_stop_words and have_curse_words:
                # Want both stop words and curse words included
                word_count_dictionary[word] = count

            elif have_stop_words and not have_curse_words:
                # Want just stop words included but curse words removed
                if word not in curse_words:
                    word_count_dictionary[word] = count

            elif not have_stop_words and have_curse_words:
                # Want stop words removed but curse words included
                if word not in stop_words:
                    word_count_dictionary[word] = count

            else:
                # Want to remove both stop and curse words
                if word not in stop_words and word not in curse_words:
                    word_count_dictionary[word] = count

        return word_count_dictionary
          
    except Exception as e:
        print(e)
        logging.basicConfig(filename='Cloud_lyrics.log', filemode='a', format='%(asctime)s - %(message)s',
                            datefmt='%b-%d-%Y %H:%M:%S')
        logging.error('Error writing word count to excel')
        logging.error('------------------')
        logging.error("Exception occurred", exc_info=True)
        logging.error('------------------')




def read_word_dictionary_json(fileIn):
    """Returns a json file that has been read in holding word count dictionary
       Args:    fileIn 
       Returns: word_count json file  
    """
    try:
        with open(fileIn) as json_file:
            word_count = json.load(json_file)

        return word_count

    except Exception as e:
        print(e)
        logging.basicConfig(filename='Cloud_lyrics.log', filemode='a', format='%(asctime)s - %(message)s',
                            datefmt='%b-%d-%Y %H:%M:%S')
        logging.error('Error reading json file')
        logging.error('------------------')
        logging.error("Exception occurred", exc_info=True)
        logging.error('------------------')


def update_json_count(previous_json_data, new_lyrics_dictionary, raw_lyrics=False):
    """Updates the running count of words by reading in the previously generated json file
       Args:    json_file, new_lyrics_dictionary, raw_lyrics
       Returns: a sorted ditionary file of the updated running word count  
    """

    for word, count in new_lyrics_dictionary.items():
        if word in previous_json_data:
            previous_json_data[word] += count
        else:
            previous_json_data[word] = count

    sorted_json = dict(sorted(previous_json_data.items(), key=lambda kv: kv[1], reverse=True))

    write_word_count_json(sorted_json, raw_lyrics)

    return sorted_json
