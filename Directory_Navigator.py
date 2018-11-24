import os
import openpyxl
import logging
import datetime
import logging
import json

ROOT_DIR = os.getcwd()

DIRECTORY_SWITCH = {

    'Spotify_Global_Charts':            ROOT_DIR + '\\' + 'Spotify_Charts\\Spotify_Global_Charts\\',
    'Spotify_US_Charts':                ROOT_DIR + '\\' + 'Spotify_Charts\\Spotify_US_Charts\\',
    'Country_Excel_Results':            ROOT_DIR + '\\' + 'Excel_Results\\Country\\',
    'Alternative_Excel_Results':        ROOT_DIR + '\\' + 'Excel_Results\\Alternative\\',
    'Metal_Excel_Results':              ROOT_DIR + '\\' + 'Excel_Results\\Metal\\',
    'Pop_Excel_Results':                ROOT_DIR + '\\' + 'Excel_Results\\Pop\\',
    'Rap_Excel_Results':                ROOT_DIR + '\\' + 'Excel_Results\\Rap\\',
    'Rock_Excel_Results':               ROOT_DIR + '\\' + 'Excel_Results\\Rock\\',
    'Spotify_Gloabl_Excel_Results':     ROOT_DIR + '\\' + 'Excel_Results\\Spotify_Global\\',
    'Spotify_US_Excel_Results':         ROOT_DIR + '\\' + 'Excel_Results\\Spotify_US\\',
    'iTunes_Excel_Results':             ROOT_DIR + '\\' + 'Excel_Results\\iTunes\\',
    'Text_File_Excel_Results':          ROOT_DIR + '\\' + 'Excel_Results\\Text Files\\',
    'Json_Filter_Files':                ROOT_DIR + '\\' + 'Json_Files\\Json_Filtered\\',
    'Json_Raw_Files':                   ROOT_DIR + '\\' + 'Json_Files\\Json_Raw\\'

}


def check_for_directories():
    """Checks to see if all directories exist and if not creates them
       Args:        -
       Returns:     -
    """
    for key, value in DIRECTORY_SWITCH.items():
        try:
            if not os.path.exists(value):
                os.makedirs(value)
            else:
                pass

        except Exception as e:
            print('Failed to make directories')


def get_existing_excel(genre):
    """Checks to see if excel files exists for selected genre
       Args:        selected genre
       Returns:     a list of all existing excel files for a selected genre
    """
    excel_directory = get_directory_excel(genre)

    excel_list = []

    for file in os.listdir(excel_directory):
        if file[-5:] == '.xlsx':

            excel_list.append(excel_directory + file)

    return excel_list

def check_for_excel(genre):
    """Checks to see if an excel file already exists for the current day for a selected genre
       Args:        selected genre
       Returns:     file path of existing excel file or None if doesnt exist
    """
    excel_files = get_existing_excel(genre)
    current_date = str(datetime.datetime.today().strftime('%m-%d-%Y'))

    for file in excel_files:
        if current_date in file:
            return file

    return None

def write_excel(word_dictionary_raw, genre, song_artist_list):
    """Writes words and thier frequncies to excel file
       Also writes the songs and their artist
       Args:        word dictionary, selected genre, song_artist list
       Returns:     file path to newly created excel 
    """
    file_out = str(datetime.datetime.today().strftime('%m-%d-%Y')) + '_' + genre + '_Excel_Data.xlsx'
    file_directory = get_directory_excel(genre)
    file_out_path = file_directory + file_out

    try:
        work_book = openpyxl.Workbook()

        sheet = work_book.active
        sheet.title = 'Word Count'

        sheet['A1'] = 'Word'
        sheet['B1'] = 'Count'

        row_counter = 2  # start counter at 2 to skip header
        column_counter = 1
        for word, count in word_dictionary_raw.items():
            sheet.cell(row=row_counter, column=column_counter).value = word
            column_counter += 1
            sheet.cell(row=row_counter, column=column_counter).value = count
            row_counter += 1
            column_counter = 1

        if song_artist_list != None:
            #Create a new sheet and set the active sheet to that newly created sheet
            work_book.create_sheet(title='Song List')

            song_list_sheet = work_book['Song List']

            song_list_sheet['A1'] = 'Song'
            song_list_sheet['B1'] = 'Artist'

            row_counter = 2
            column_counter = 1
            for song_artist_pair in song_artist_list:
                song_list_sheet.cell(row=row_counter, column=column_counter).value = song_artist_pair[0]
                column_counter += 1
                song_list_sheet.cell(row=row_counter, column=column_counter).value = song_artist_pair[1]
                row_counter += 1
                column_counter = 1


        work_book.save(file_out_path)

        return file_out_path

    except Exception as e:
        print(e)
        logging.basicConfig(filename='Cloud_lyrics.log', filemode='a', format='%(asctime)s - %(message)s',
                            datefmt='%b-%d-%Y %H:%M:%S')
        logging.error('Error writing word count to excel')
        logging.error('------------------')
        logging.error("Exception occurred", exc_info=True)
        logging.error('------------------')


def write_csv(file_type, file_content):
    """Saves spotify charts csv file
       Args:        file tpye, file content
       Returns:     file out path 
    """
    file_out = str(datetime.datetime.today().strftime('%m-%d-%Y')) + '_' + file_type + '.csv'
    file_directory = get_directory_charts(file_type)
    file_out_path = file_directory + file_out

    open(file_out_path, 'wb').write(file_content)

    return file_out_path


def write_json(raw_json, file_content):
    """Writes words and their frequencies to json file
       Args:        raw json boolean, file content
       Returns:     file path  
    """
    if raw_json:
        file_out = str(datetime.datetime.today().strftime('%m-%d-%Y')) + '_RAW_JSON_Data.json'
        file_directory = get_directory_json('Raw')
        file_out_path = file_directory + file_out

        with open(file_out_path, 'w') as json_file:
            json.dump(file_content, json_file)

        return file_out_path

    else:
        file_out = str(datetime.datetime.today().strftime('%m-%d-%Y')) + '_JSON_Data.json'
        file_directory = get_directory_json('Filtered')
        file_out_path = file_directory + file_out

        with open(file_out_path, 'w') as json_file:
            json.dump(file_content, json_file)

        return file_out_path


def get_directory_json(category):
    """Returns a directory for the selected type
       Args:        category
       Returns:     directory 
    """
    if 'Raw' in category:
        return DIRECTORY_SWITCH['Json_Raw_Files']

    elif 'Filtered' in category:
        return DIRECTORY_SWITCH['Json_Filter_Files']


def get_directory_excel(category):
    """Returns a directory for the selected type
       Args:        category
       Returns:     directory
    """
    if 'Country' == category:
        return DIRECTORY_SWITCH['Country_Excel_Results']

    elif 'Metal' == category:
        return DIRECTORY_SWITCH['Metal_Excel_Results']

    elif 'Alternative' == category:
        return DIRECTORY_SWITCH['Alternative_Excel_Results']

    elif 'Pop' == category:
        return DIRECTORY_SWITCH['Pop_Excel_Results']

    elif 'Rap' == category:
        return DIRECTORY_SWITCH['Rap_Excel_Results']

    elif 'Rock' == category:
        return DIRECTORY_SWITCH['Rock_Excel_Results']

    elif 'iTunes' == category:
        return DIRECTORY_SWITCH['iTunes_Excel_Results']

    elif 'Spotify_Global' == category:
        return DIRECTORY_SWITCH['Spotify_Gloabl_Excel_Results']

    elif 'Spotify_US' == category:
        return DIRECTORY_SWITCH['Spotify_US_Excel_Results']

    elif 'Text_File' == category:
        return DIRECTORY_SWITCH['Text_File_Excel_Results']


def get_directory_charts(category):
    """Returns a directory to the spoitify csv charts
       Args:        category
       Returns:     directory 
    """    
    if 'Global' in category:
        return DIRECTORY_SWITCH['Spotify_Global_Charts']

    elif 'US' in category:
        return DIRECTORY_SWITCH['Spotify_US_Charts']
